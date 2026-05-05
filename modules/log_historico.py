"""
log_historico.py

Mantém um log histórico de todas as execuções do script em um arquivo Excel
salvo em: <pasta_raiz_projeto>/log/log_execucoes.xlsx

Colunas do Excel:
    - execucao_id      : identificador único da execução (timestamp)
    - data_inicio      : data/hora de início
    - data_fim         : data/hora de fim
    - duracao_total    : tempo total formatado (ex: 1h 23m 45s)
    - competencia      : pasta/competência processada
    - total_apis       : quantas APIs foram tentadas
    - apis_sucesso     : quantas tiveram sucesso
    - apis_erro        : quantas falharam
    - total_registros  : soma de registros extraídos
    - modo             : 'processar' ou 'copiar'
    - status_final     : 'Sucesso Total', 'Parcial' ou 'Falha'
    - apis_detalhes    : lista resumida das APIs e seus status
    - observacoes      : qualquer nota extra
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False
    print("⚠️  openpyxl não encontrado. Instale com: pip install openpyxl")


# Cores
COR_CABECALHO = "1F3864"
COR_SUCESSO   = "C6EFCE"
COR_PARCIAL   = "FFEB9C"
COR_FALHA     = "FFC7CE"
COR_LINHA_PAR = "F2F2F2"

LOG_FOLDER   = "log"
LOG_FILENAME = "log_execucoes.xlsx"

COLUNAS = [
    ("execucao_id",     "ID Execução",         22),
    ("data_inicio",     "Início",              20),
    ("data_fim",        "Fim",                 20),
    ("duracao_total",   "Duração",             14),
    ("competencia",     "Competência",         30),
    ("total_apis",      "Total APIs",          12),
    ("apis_sucesso",    "Sucesso",             12),
    ("apis_erro",       "Erros",               12),
    ("total_registros", "Registros",           14),
    ("modo",            "Modo",                14),
    ("status_final",    "Status",              16),
    ("apis_detalhes",   "Detalhe das APIs",    60),
    ("observacoes",     "Observações",         40),
]


def _caminho_log(pasta_raiz_projeto: str) -> str:
    """Retorna o caminho completo do log dentro da pasta log/."""
    pasta_log = os.path.join(pasta_raiz_projeto, LOG_FOLDER)
    os.makedirs(pasta_log, exist_ok=True)   # cria a pasta se não existir
    return os.path.join(pasta_log, LOG_FILENAME)


class LogHistorico:
    """
    Gerencia o log histórico de execuções em Excel.
    Cada chamada a `fechar_execucao()` adiciona uma nova linha ao arquivo.

    Uso:
        log = LogHistorico(pasta_raiz_projeto=PASTA_RAIZ_PROJETO)
        log.iniciar_execucao(competencia=caminho, modo="processar")
        ...
        resumo = log.fechar_execucao(resultados=resultados,
                                     total_registros=1234)
    """

    def __init__(self, pasta_raiz_projeto: str):
        if not OPENPYXL_DISPONIVEL:
            raise ImportError("openpyxl é necessário. Instale com: pip install openpyxl")

        self.caminho      = _caminho_log(pasta_raiz_projeto)
        self._inicio      = None
        self._execucao_id = ""
        self._competencia = ""
        self._modo        = ""

        print(f"📂 Arquivo de log: {self.caminho}")

    def iniciar_execucao(self, competencia: str = "", modo: str = "processar") -> str:
        """
        Marca o início de uma execução.

        Args:
            competencia: Pasta/competência que está sendo processada.
            modo:        'processar' ou 'copiar'.

        Returns:
            execucao_id (string com timestamp).
        """
        self._inicio      = datetime.now()
        self._execucao_id = self._inicio.strftime("EXE_%Y%m%d_%H%M%S")
        self._competencia = competencia
        self._modo        = modo

        print(f"📋 Log histórico iniciado → ID: {self._execucao_id}")
        return self._execucao_id

    def fechar_execucao(
        self,
        resultados: dict,
        total_registros: int = 0,
        observacoes: str = "",
    ) -> dict:
        """
        Fecha a execução, persiste a linha no Excel e retorna o dicionário-resumo
        para ser usado no envio de e-mail.

        Args:
            resultados:      Dict {nome_api: {'sucesso': bool, 'arquivo': ..., 'erro': ...}}
            total_registros: Total de registros extraídos na execução.
            observacoes:     Texto livre opcional.

        Returns:
            Dicionário resumo pronto para ser passado a enviar_email_extracao().
        """
        fim    = datetime.now()
        inicio = self._inicio or fim

        # Métricas
        total_apis = len(resultados)
        apis_ok    = sum(1 for r in resultados.values() if r.get('sucesso'))
        apis_erro  = total_apis - apis_ok
        duracao    = self._formatar_duracao(inicio, fim)

        if total_apis == 0:
            status = "Falha"
        elif apis_ok == total_apis:
            status = "Sucesso Total"
        elif apis_ok > 0:
            status = "Parcial"
        else:
            status = "Falha"

        # Detalhe resumido das APIs para o Excel
        detalhes_str = " | ".join(
            f"{'✅' if r.get('sucesso') else '❌'} {nome}"
            for nome, r in resultados.items()
        )

        # Detalhe por API em formato dict para o e-mail
        apis_detalhes_dict = {
            nome: bool(r.get('sucesso')) for nome, r in resultados.items()
        }

        linha = {
            "execucao_id":      self._execucao_id,
            "data_inicio":      inicio.strftime('%d/%m/%Y %H:%M:%S'),
            "data_fim":         fim.strftime('%d/%m/%Y %H:%M:%S'),
            "duracao_total":    duracao,
            "competencia":      self._competencia,
            "total_apis":       total_apis,
            "apis_sucesso":     apis_ok,
            "apis_erro":        apis_erro,
            "total_registros":  total_registros,
            "modo":             self._modo,
            "status_final":     status,
            "apis_detalhes":    detalhes_str,
            "observacoes":      observacoes,
        }

        self._salvar_linha(linha)

        print(f"📋 Log histórico fechado  → Status: {status} | Duração: {duracao}")

        # Resumo para o e-mail — inclui todos os campos que enviar_email_extracao() espera
        resumo = {
            "modo":                    self._modo,
            "status_final":            status,
            "total_apis":              total_apis,
            "apis_sucesso":            apis_ok,
            "apis_erro":               apis_erro,
            "duracao_total":           duracao,
            "total_registros":         total_registros,
            "data_inicio":             inicio.strftime('%d/%m/%Y %H:%M:%S'),
            "data_fim":                fim.strftime('%d/%m/%Y %H:%M:%S'),
            "competencias_processadas": [],   # preenchido pelo main.py se disponível
            "apis_detalhes":           apis_detalhes_dict,
            "observacoes":             observacoes,
            "caminho_log":             self.caminho,
        }

        return resumo

    def _salvar_linha(self, linha: dict) -> None:
        if os.path.exists(self.caminho):
            wb = openpyxl.load_workbook(self.caminho)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Execuções"
            self._criar_cabecalho(ws)

        status = linha.get("status_final", "")
        if status == "Sucesso Total":
            cor_linha = COR_SUCESSO
        elif status == "Parcial":
            cor_linha = COR_PARCIAL
        elif status == "Falha":
            cor_linha = COR_FALHA
        else:
            cor_linha = None

        num_linha  = ws.max_row + 1
        usar_zebra = (num_linha % 2 == 0) and cor_linha is None

        valores = [linha.get(col[0], "") for col in COLUNAS]
        ws.append(valores)

        row_idx = ws.max_row
        fill = (PatternFill("solid", fgColor=cor_linha) if cor_linha else
                PatternFill("solid", fgColor=COR_LINHA_PAR) if usar_zebra else None)

        for col_idx, (campo, _, _largura) in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                wrap_text=(campo == "apis_detalhes"),
                vertical="center"
            )
            if fill:
                cell.fill = fill
            self._aplicar_borda_fina(cell)

        ws.row_dimensions[row_idx].height = 30

        try:
            wb.save(self.caminho)
            print(f"💾 Log salvo em: {self.caminho}")
        except PermissionError:
            alt = self.caminho.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}.xlsx")
            wb.save(alt)
            print(f"⚠️  Arquivo em uso. Log salvo em: {alt}")
            self.caminho = alt

    def _criar_cabecalho(self, ws) -> None:
        cabecalhos = [col[1] for col in COLUNAS]
        ws.append(cabecalhos)

        fill_header = PatternFill("solid", fgColor=COR_CABECALHO)
        font_header = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, (campo, label, largura) in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = fill_header
            cell.font      = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            self._aplicar_borda_fina(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    @staticmethod
    def _aplicar_borda_fina(cell) -> None:
        lado = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=lado, right=lado, top=lado, bottom=lado)

    @staticmethod
    def _formatar_duracao(inicio: datetime, fim: datetime) -> str:
        total    = int((fim - inicio).total_seconds())
        horas    = total // 3600
        minutos  = (total % 3600) // 60
        segundos = total % 60
        if horas > 0:
            return f"{horas}h {minutos:02d}m {segundos:02d}s"
        elif minutos > 0:
            return f"{minutos}m {segundos:02d}s"
        else:
            return f"{segundos}s"